import openpyxl
from openpyxl.styles import (
    Font, Alignment, PatternFill, Border, Side
)
from openpyxl.utils import get_column_letter

# ─── Color palette ───────────────────────────────────────────────
C_HEADER_BG   = "1F4E79"   # dark navy
C_HEADER_FONT = "FFFFFF"
C_SUB_BG      = "BDD7EE"   # light blue
C_ALT_ROW     = "EBF3FB"
C_LABEL_BG    = "D6E4F0"
C_BORDER      = "2F75B6"
C_TOTAL_BG    = "FCE4D6"   # light orange
C_PLATFORM_TT = "FF0050"   # TikTok red
C_PLATFORM_SP = "EE4D2D"   # Shopee orange

def thin_border(color="2F75B6"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def medium_border():
    s = Side(style="medium", color="1F4E79")
    return Border(left=s, right=s, top=s, bottom=s)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def bold_font(size=10, color="000000"):
    return Font(name="Arial", bold=True, size=size, color=color)

def normal_font(size=10, color="000000"):
    return Font(name="Arial", size=size, color=color)

def center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def left(wrap=False):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

def right():
    return Alignment(horizontal="right", vertical="center")

def set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width

def merge_write(ws, cell_range, value, font=None, align=None, fill_=None, border=None):
    ws.merge_cells(cell_range)
    cell = ws[cell_range.split(":")[0]]
    cell.value = value
    if font:   cell.font   = font
    if align:  cell.alignment = align
    if fill_:  cell.fill   = fill_
    if border: cell.border = border

def apply_border_range(ws, min_row, max_row, min_col, max_col, border=None):
    if border is None:
        border = thin_border()
    for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                             min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = border

# ══════════════════════════════════════════════════════════════════
#  SHEET 1 – PHIẾU XUẤT KHO
# ══════════════════════════════════════════════════════════════════
def build_phieu_xuat_kho(wb):
    ws = wb.create_sheet("PHIẾU XUẤT KHO")
    ws.sheet_view.showGridLines = False

    # column widths
    widths = {"A":5,"B":18,"C":32,"D":12,"E":12,"F":15,"G":15,"H":15,"I":18,"J":20}
    for col, w in widths.items():
        set_col_width(ws, col, w)

    ws.row_dimensions[1].height = 8

    # ── Logo / tiêu đề công ty ──────────────────────────────────
    merge_write(ws, "B2:D5",
        "CÔNG TY / CỬA HÀNG\n[TÊN ĐƠN VỊ]",
        font=Font(name="Arial", bold=True, size=13, color="1F4E79"),
        align=Alignment(horizontal="left", vertical="center", wrap_text=True))

    merge_write(ws, "E2:J2",
        "PHIẾU XUẤT KHO",
        font=Font(name="Arial", bold=True, size=18, color=C_HEADER_FONT),
        align=center(),
        fill_=fill(C_HEADER_BG))
    ws.row_dimensions[2].height = 30

    merge_write(ws, "E3:J3",
        "THƯƠNG MẠI ĐIỆN TỬ – TIKTOK SHOP / SHOPEE",
        font=Font(name="Arial", bold=True, size=10, color=C_HEADER_FONT),
        align=center(),
        fill_=fill("2E75B6"))
    ws.row_dimensions[3].height = 18

    merge_write(ws, "E4:G4", "Số phiếu (No.):",
        font=bold_font(10), align=left(), fill_=fill(C_LABEL_BG))
    merge_write(ws, "H4:J4", "PXK-2026-_____",
        font=Font(name="Arial", size=10, color="C00000", bold=True), align=center())
    apply_border_range(ws, 4, 4, 5, 10, thin_border())

    merge_write(ws, "E5:G5", "Ngày xuất kho:",
        font=bold_font(10), align=left(), fill_=fill(C_LABEL_BG))
    merge_write(ws, "H5:J5", "___/___/20____",
        font=normal_font(10), align=center())
    apply_border_range(ws, 5, 5, 5, 10, thin_border())

    ws.row_dimensions[6].height = 6

    # ── Thông tin chung ─────────────────────────────────────────
    info_rows = [
        ("B7:C7", "Sàn TMĐT:",          "D7:J7",  "☐ TikTok Shop    ☐ Shopee    ☐ Lazada    ☐ Khác: ___________"),
        ("B8:C8", "Mã đơn hàng (Order ID):", "D8:J8", ""),
        ("B9:C9", "Người nhận hàng / Shipper:", "D9:J9", ""),
        ("B10:C10","Đơn vị vận chuyển:", "D10:J10","☐ GHN   ☐ GHTK   ☐ J&T   ☐ Viettel Post   ☐ Ninja Van   ☐ Khác"),
        ("B11:C11","Mã vận đơn (Tracking):", "D11:J11",""),
        ("B12:C12","Kho xuất:",           "D12:J12",""),
        ("B13:C13","Người xuất kho:",     "D13:J13",""),
    ]
    for label_range, label_val, val_range, val_val in info_rows:
        merge_write(ws, label_range, label_val,
                    font=bold_font(9), align=left(), fill_=fill(C_LABEL_BG))
        merge_write(ws, val_range, val_val,
                    font=normal_font(9), align=left())
        start_cell = label_range.split(":")[0]
        row_num = int(''.join(c for c in start_cell if c.isdigit()))
        apply_border_range(ws, row_num, row_num, 2, 10, thin_border())

    ws.row_dimensions[14].height = 6

    # ── Bảng sản phẩm header ────────────────────────────────────
    headers = ["STT","Mã SKU","Tên sản phẩm / Mô tả","Đơn vị","Màu / Size","Số lượng\nYêu cầu","Số lượng\nThực xuất","Đơn giá\n(VNĐ)","Thành tiền\n(VNĐ)","Ghi chú"]
    for col_idx, h in enumerate(headers, start=2):
        cell = ws.cell(row=15, column=col_idx, value=h)
        cell.font = Font(name="Arial", bold=True, size=9, color=C_HEADER_FONT)
        cell.alignment = center(wrap=True)
        cell.fill = fill(C_HEADER_BG)
        cell.border = thin_border()
    ws.row_dimensions[15].height = 36

    # ── Data rows (10 rows) ─────────────────────────────────────
    for r in range(16, 26):
        ws.row_dimensions[r].height = 22
        bg = fill(C_ALT_ROW) if r % 2 == 0 else fill("FFFFFF")
        for col_idx in range(2, 12):
            cell = ws.cell(row=r, column=col_idx)
            if col_idx == 2:
                cell.value = r - 15
                cell.alignment = center()
                cell.font = bold_font(9)
            else:
                cell.alignment = center(wrap=True) if col_idx in (4,5) else left()
                cell.font = normal_font(9)
            cell.fill = bg
            cell.border = thin_border()
            if col_idx in (8, 9, 10):
                cell.number_format = '#,##0'

    ws.row_dimensions[26].height = 8

    # ── Tổng cộng ───────────────────────────────────────────────
    merge_write(ws, "B27:G27", "TỔNG CỘNG",
        font=bold_font(10, "FFFFFF"), align=center(), fill_=fill(C_HEADER_BG))
    merge_write(ws, "H27:H27", "=SUM(H16:H25)",
        font=bold_font(10), align=center(), fill_=fill(C_TOTAL_BG))
    ws["H27"].number_format = '#,##0'
    merge_write(ws, "I27:I27", "=SUM(I16:I25)",
        font=bold_font(10), align=right(), fill_=fill(C_TOTAL_BG))
    ws["I27"].number_format = '#,##0'
    merge_write(ws, "J27:J27", "",
        font=normal_font(9), fill_=fill(C_TOTAL_BG))
    apply_border_range(ws, 27, 27, 2, 10, medium_border())

    # ── Số tiền bằng chữ ────────────────────────────────────────
    merge_write(ws, "B28:C28", "Bằng chữ:",
        font=bold_font(9), align=left(), fill_=fill(C_LABEL_BG))
    merge_write(ws, "D28:J28", "",
        font=Font(name="Arial", size=9, italic=True), align=left())
    apply_border_range(ws, 28, 28, 2, 10, thin_border())
    ws.row_dimensions[28].height = 18

    # ── Ghi chú ─────────────────────────────────────────────────
    merge_write(ws, "B29:C29", "Ghi chú đặc biệt:",
        font=bold_font(9), align=left(), fill_=fill(C_LABEL_BG))
    merge_write(ws, "D29:J29", "",
        font=normal_font(9), align=left())
    apply_border_range(ws, 29, 29, 2, 10, thin_border())
    ws.row_dimensions[29].height = 18

    ws.row_dimensions[30].height = 14

    # ── Ký tên ──────────────────────────────────────────────────
    sig_sections = [
        ("B31:C31", "NGƯỜI LẬP PHIẾU"),
        ("D31:E31", "THỦ KHO"),
        ("F31:G31", "KẾ TOÁN"),
        ("H31:J31", "NGƯỜI NHẬN / SHIPPER"),
    ]
    for cell_range, title in sig_sections:
        merge_write(ws, cell_range, title,
            font=bold_font(9, C_HEADER_FONT), align=center(),
            fill_=fill(C_HEADER_BG))
    apply_border_range(ws, 31, 31, 2, 10, thin_border())
    ws.row_dimensions[31].height = 22

    for r in range(32, 36):
        ws.row_dimensions[r].height = 18
        for col_group in [(2,3),(4,5),(6,7),(8,10)]:
            for c in range(col_group[0], col_group[1]+1):
                ws.cell(row=r, column=c).border = thin_border()

    sub_labels = [
        ("B35:C35", "(Ký, ghi rõ họ tên)"),
        ("D35:E35", "(Ký, ghi rõ họ tên)"),
        ("F35:G35", "(Ký, ghi rõ họ tên)"),
        ("H35:J35", "(Ký, ghi rõ họ tên – Mã nhân viên giao hàng)"),
    ]
    for cell_range, lbl in sub_labels:
        merge_write(ws, cell_range, lbl,
            font=Font(name="Arial", size=8, italic=True, color="595959"),
            align=center())
        apply_border_range(ws, 35, 35, 2, 10, thin_border())

    # ── Outer border toàn bộ ────────────────────────────────────
    apply_border_range(ws, 2, 35, 2, 10, thin_border())
    ws.print_area = "B2:J36"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = 9  # A4
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1


# ══════════════════════════════════════════════════════════════════
#  SHEET 2 – BIÊN BẢN XÁC NHẬN GIAO HÀNG VỚI SHIPPER
# ══════════════════════════════════════════════════════════════════
def build_bien_ban_giao_hang(wb):
    ws = wb.create_sheet("BIÊN BẢN GIAO HÀNG SHIPPER")
    ws.sheet_view.showGridLines = False

    widths = {
        "A":4, "B":5, "C":22, "D":22, "E":15,
        "F":12, "G":12, "H":15, "I":15, "J":12,
        "K":16, "L":18
    }
    for col, w in widths.items():
        set_col_width(ws, col, w)

    ws.row_dimensions[1].height = 8

    # ── Tiêu đề ─────────────────────────────────────────────────
    merge_write(ws, "B2:D5",
        "CÔNG TY / CỬA HÀNG\n[TÊN ĐƠN VỊ]\nĐịa chỉ: _______________\nĐT: _______________",
        font=Font(name="Arial", bold=True, size=10, color="1F4E79"),
        align=Alignment(horizontal="left", vertical="center", wrap_text=True))
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 20
    ws.row_dimensions[5].height = 20

    merge_write(ws, "E2:L3",
        "BIÊN BẢN XÁC NHẬN GIAO HÀNG VỚI SHIPPER",
        font=Font(name="Arial", bold=True, size=16, color=C_HEADER_FONT),
        align=center(),
        fill_=fill(C_HEADER_BG))

    merge_write(ws, "E4:L4",
        "THƯƠNG MẠI ĐIỆN TỬ – TIKTOK SHOP / SHOPEE",
        font=Font(name="Arial", bold=True, size=10, color=C_HEADER_FONT),
        align=center(),
        fill_=fill("2E75B6"))

    merge_write(ws, "E5:H5", "Số biên bản:",
        font=bold_font(10), align=left(), fill_=fill(C_LABEL_BG))
    merge_write(ws, "I5:L5", "BB-GH-2026-_____",
        font=Font(name="Arial", size=10, bold=True, color="C00000"), align=center())
    apply_border_range(ws, 5, 5, 5, 12, thin_border())

    ws.row_dimensions[6].height = 6

    # ── Thông tin bàn giao ──────────────────────────────────────
    info_rows2 = [
        ("B7:D7",  "Ngày bàn giao:",             "E7:L7",  "___/___/20____     Giờ bàn giao: ____:____"),
        ("B8:D8",  "Đơn vị vận chuyển (Shipper):","E8:L8",  "☐ GHN   ☐ GHTK   ☐ J&T Express   ☐ Viettel Post   ☐ Ninja Van   ☐ Khác: ______"),
        ("B9:D9",  "Mã nhân viên / Tên Shipper:", "E9:L9",  ""),
        ("B10:D10","SĐT Shipper:",                 "E10:H10",""),
        ("E10:H10","",                              None,None),
        ("I10:L10","Biển số xe:",                  None,None),
        ("B11:D11","Người bàn giao (Kho):",        "E11:L11",""),
        ("B12:D12","Sàn TMĐT:",                    "E12:L12","☐ TikTok Shop   ☐ Shopee   ☐ Cả hai   ☐ Khác: ______"),
        ("B13:D13","Tổng số kiện hàng:",           "E13:H13",""),
        ("B14:D14","Ghi chú chung:",               "E14:L14",""),
    ]

    # Simplified info row writing
    simple_rows = [
        ("B7:D7",  "Ngày bàn giao:",              "E7:L7",   "___/___/20____     Giờ bàn giao: ____:____"),
        ("B8:D8",  "Đơn vị vận chuyển (Shipper):","E8:L8",   "☐ GHN   ☐ GHTK   ☐ J&T Express   ☐ Viettel Post   ☐ Ninja Van   ☐ Khác"),
        ("B9:D9",  "Mã NV / Tên Shipper:",        "E9:L9",   ""),
        ("B11:D11","Người bàn giao (Kho):",        "E11:L11", ""),
        ("B12:D12","Sàn TMĐT:",                    "E12:L12", "☐ TikTok Shop   ☐ Shopee   ☐ Cả hai   ☐ Khác: ______"),
        ("B14:D14","Ghi chú chung:",               "E14:L14", ""),
    ]
    for label_range, label_val, val_range, val_val in simple_rows:
        merge_write(ws, label_range, label_val,
                    font=bold_font(9), align=left(), fill_=fill(C_LABEL_BG))
        merge_write(ws, val_range, val_val,
                    font=normal_font(9), align=left())
        r = int(label_range.split(":")[0][1:])
        apply_border_range(ws, r, r, 2, 12, thin_border())

    # Row 10 split
    merge_write(ws, "B10:D10", "SĐT Shipper:",
        font=bold_font(9), align=left(), fill_=fill(C_LABEL_BG))
    merge_write(ws, "E10:H10", "",
        font=normal_font(9), align=left())
    merge_write(ws, "I10:J10", "Biển số xe:",
        font=bold_font(9), align=left(), fill_=fill(C_LABEL_BG))
    merge_write(ws, "K10:L10", "",
        font=normal_font(9), align=left())
    apply_border_range(ws, 10, 10, 2, 12, thin_border())

    # Row 13 split
    merge_write(ws, "B13:D13", "Tổng số kiện hàng:",
        font=bold_font(9), align=left(), fill_=fill(C_LABEL_BG))
    merge_write(ws, "E13:H13", "",
        font=normal_font(9), align=left())
    merge_write(ws, "I13:J13", "Tổng COD (VNĐ):",
        font=bold_font(9), align=left(), fill_=fill(C_LABEL_BG))
    merge_write(ws, "K13:L13", "",
        font=Font(name="Arial", size=9, bold=True, color="C00000"), align=right())
    apply_border_range(ws, 13, 13, 2, 12, thin_border())
    ws["K13"].number_format = '#,##0'

    ws.row_dimensions[15].height = 8

    # ── Bảng danh sách đơn hàng ─────────────────────────────────
    col_headers = [
        "STT", "Mã đơn hàng\n(Order ID)", "Mã vận đơn\n(Tracking No.)",
        "Sàn\nTMĐT", "Tên KH /\nĐịa chỉ giao", "SĐT\nKhách",
        "Số\nkiện", "COD\n(VNĐ)", "Ghi chú\nhàng hóa",
        "Trạng thái\nbàn giao", "Ký nhận\nShipper"
    ]
    for col_idx, h in enumerate(col_headers, start=2):
        cell = ws.cell(row=16, column=col_idx, value=h)
        cell.font = Font(name="Arial", bold=True, size=9, color=C_HEADER_FONT)
        cell.alignment = center(wrap=True)
        cell.fill = fill(C_HEADER_BG)
        cell.border = thin_border()
    ws.row_dimensions[16].height = 40

    # ── Data rows (15 rows) ─────────────────────────────────────
    platform_opts = "TikTok / Shopee"
    for r in range(17, 32):
        ws.row_dimensions[r].height = 22
        bg = fill(C_ALT_ROW) if r % 2 == 0 else fill("FFFFFF")
        for col_idx in range(2, 13):
            cell = ws.cell(row=r, column=col_idx)
            if col_idx == 2:
                cell.value = r - 16
                cell.alignment = center()
                cell.font = bold_font(9)
            elif col_idx == 5:
                cell.alignment = center(wrap=True)
                cell.font = normal_font(9)
            elif col_idx == 9:
                cell.alignment = left(wrap=True)
                cell.font = normal_font(9)
            elif col_idx == 11:
                cell.alignment = right()
                cell.number_format = '#,##0'
                cell.font = normal_font(9)
            elif col_idx == 13:
                cell.alignment = left(wrap=True)
                cell.font = normal_font(9)
            else:
                cell.alignment = center()
                cell.font = normal_font(9)
            cell.fill = bg
            cell.border = thin_border()

    ws.row_dimensions[32].height = 8

    # ── Tổng kết ────────────────────────────────────────────────
    merge_write(ws, "B33:G33", "TỔNG CỘNG",
        font=bold_font(10, C_HEADER_FONT), align=center(), fill_=fill(C_HEADER_BG))
    merge_write(ws, "H33:H33", "=SUM(H17:H31)",
        font=bold_font(10), align=center(), fill_=fill(C_TOTAL_BG))
    ws["H33"].number_format = '#,##0'
    merge_write(ws, "I33:I33", "=SUM(I17:I31)",
        font=bold_font(10), align=right(), fill_=fill(C_TOTAL_BG))
    ws["I33"].number_format = '#,##0'
    # Actually columns: B=STT, C=OrderID, D=Tracking, E=Sàn, F=TênKH, G=SĐT, H=SốKiện, I=COD, J=GhiChú, K=TT, L=Ký
    merge_write(ws, "J33:L33", "",
        font=normal_font(9), fill_=fill(C_TOTAL_BG))
    apply_border_range(ws, 33, 33, 2, 12, medium_border())

    ws.row_dimensions[34].height = 10

    # ── Điều khoản bàn giao ─────────────────────────────────────
    merge_write(ws, "B35:L35",
        "ĐIỀU KHOẢN BÀN GIAO",
        font=bold_font(10, C_HEADER_FONT), align=center(), fill_=fill("2E75B6"))
    ws.row_dimensions[35].height = 20

    terms = [
        "1. Shipper có trách nhiệm kiểm tra số lượng kiện hàng trước khi ký nhận. Sau khi ký nhận, shipper chịu trách nhiệm về hàng hóa.",
        "2. Hàng hóa phải được giao đúng địa chỉ ghi trên vận đơn. Mọi thắc mắc liên hệ kho trong vòng 24 giờ.",
        "3. Trường hợp hàng bị hỏng / mất, shipper chịu trách nhiệm bồi thường theo quy định của đơn vị vận chuyển.",
        "4. COD phải được nộp về kho/kế toán trong vòng 24-48h sau khi giao thành công.",
    ]
    for i, term in enumerate(terms):
        r = 36 + i
        merge_write(ws, f"B{r}:L{r}", term,
            font=Font(name="Arial", size=8, italic=True),
            align=Alignment(horizontal="left", vertical="center", wrap_text=True))
        ws.row_dimensions[r].height = 18
        apply_border_range(ws, r, r, 2, 12, thin_border())

    ws.row_dimensions[40].height = 10

    # ── Ký tên ──────────────────────────────────────────────────
    sig_cols = [
        ("B41:D41", "ĐẠI DIỆN KHO\n(Người bàn giao)"),
        ("E41:G41", "KẾ TOÁN\n(Xác nhận COD)"),
        ("H41:J41", "SHIPPER\n(Ký nhận hàng)"),
        ("K41:L41", "QUẢN LÝ\n(Duyệt)"),
    ]
    for cell_range, title in sig_cols:
        merge_write(ws, cell_range, title,
            font=bold_font(9, C_HEADER_FONT), align=center(wrap=True),
            fill_=fill(C_HEADER_BG))
    ws.row_dimensions[41].height = 30

    for r in range(42, 47):
        ws.row_dimensions[r].height = 18
        apply_border_range(ws, r, r, 2, 12, thin_border())

    sub_labels2 = [
        ("B46:D46", "(Ký, ghi rõ họ tên)"),
        ("E46:G46", "(Ký, ghi rõ họ tên)"),
        ("H46:J46", "(Ký tên + Mã NV shipper)"),
        ("K46:L46", "(Ký, ghi rõ họ tên)"),
    ]
    for cell_range, lbl in sub_labels2:
        merge_write(ws, cell_range, lbl,
            font=Font(name="Arial", size=8, italic=True, color="595959"),
            align=center())
    apply_border_range(ws, 46, 46, 2, 12, thin_border())

    ws.print_area = "B2:L47"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = 9
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1


# ══════════════════════════════════════════════════════════════════
#  SHEET 3 – HƯỚNG DẪN SỬ DỤNG
# ══════════════════════════════════════════════════════════════════
def build_huong_dan(wb):
    ws = wb.create_sheet("HƯỚNG DẪN SỬ DỤNG")
    ws.sheet_view.showGridLines = False
    set_col_width(ws, "A", 3)
    set_col_width(ws, "B", 28)
    set_col_width(ws, "C", 60)

    merge_write(ws, "B1:C1", "HƯỚNG DẪN SỬ DỤNG MẪU EXCEL",
        font=Font(name="Arial", bold=True, size=14, color=C_HEADER_FONT),
        align=center(), fill_=fill(C_HEADER_BG))
    ws.row_dimensions[1].height = 30

    guide_items = [
        ("PHIẾU XUẤT KHO", None),
        ("Số phiếu (PXK-YYYY-xxxxx)", "Đặt theo quy tắc: PXK-NĂM-SỐ THỨTỰ (vd: PXK-2026-00001)"),
        ("Sàn TMĐT", "Đánh dấu ☑ vào sàn tương ứng: TikTok Shop hoặc Shopee"),
        ("Mã đơn hàng", "Copy mã đơn từ TikTok Shop / Shopee (Order ID)"),
        ("Mã vận đơn", "Điền sau khi sàn cấp tracking (có thể để trống lúc xuất kho)"),
        ("Số lượng yêu cầu vs thực xuất", "Ghi rõ 2 cột để phát hiện sai lệch tồn kho"),
        ("BIÊN BẢN GIAO HÀNG SHIPPER", None),
        ("Số biên bản (BB-GH-YYYY-xxxxx)", "Đặt theo quy tắc: BB-GH-NĂM-SỐ THỨ TỰ"),
        ("COD", "Ghi tổng số tiền thu hộ (Cash On Delivery) cần thu từ khách"),
        ("Trạng thái bàn giao", "Ghi: Đã bàn giao / Chờ lấy / Không lấy / Hoãn"),
        ("Ký nhận Shipper", "Shipper ký vào ô này sau khi nhận kiện hàng tương ứng"),
        ("LƯU Ý CHUNG", None),
        ("Lưu trữ", "In và lưu bản giấy. File Excel lưu tại thư mục theo tháng/năm"),
        ("Đối chiếu", "Cuối ngày đối chiếu Phiếu Xuất Kho vs Biên Bản Giao Hàng"),
        ("Tích hợp", "Có thể liên kết với báo cáo tồn kho và báo cáo doanh thu"),
    ]

    row = 2
    for item in guide_items:
        ws.row_dimensions[row].height = 22
        if item[1] is None:
            merge_write(ws, f"B{row}:C{row}", item[0],
                font=bold_font(11, C_HEADER_FONT), align=left(),
                fill_=fill("2E75B6"))
        else:
            ws.cell(row=row, column=2, value=item[0]).font = bold_font(9)
            ws.cell(row=row, column=2).fill = fill(C_LABEL_BG)
            ws.cell(row=row, column=2).alignment = left()
            ws.cell(row=row, column=2).border = thin_border()
            ws.cell(row=row, column=3, value=item[1]).font = normal_font(9)
            ws.cell(row=row, column=3).alignment = left(wrap=True)
            ws.cell(row=row, column=3).border = thin_border()
        row += 1


# ══════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()
wb.remove(wb.active)  # remove default sheet

build_phieu_xuat_kho(wb)
build_bien_ban_giao_hang(wb)
build_huong_dan(wb)

output_path = "/home/user/luuthienan/Mau_Xuat_Kho_Giao_Hang_TMDT.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")
