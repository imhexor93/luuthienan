"""
Script tạo file Excel Master Dashboard cho chiến dịch ra mắt TIAN
(Lọ tỏa hương & Lọ treo bồn cầu) - 6 Giai đoạn.
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.formatting.rule import CellIsRule, DataBarRule
from datetime import datetime, timedelta

# ============================================================
# 1. Định nghĩa dữ liệu Master Dashboard
# ============================================================

# Ngày Launch giả định (D-Day)
D_DAY = datetime(2026, 5, 15)

def d(offset):
    """Trả về ngày dựa trên offset so với D-Day."""
    return (D_DAY + timedelta(days=offset)).strftime("%Y-%m-%d")

tasks = [
    # ── Giai đoạn 1: Lên Concept & Chuẩn bị (D-30 → D-21) ──
    {
        "Mã Task": "GD1-01",
        "Giai đoạn": "GĐ1: Lên Concept & Chuẩn bị",
        "Phòng ban": "Marketing",
        "Hạng mục công việc": "Chốt USP sản phẩm (Lọ tỏa hương & Lọ treo bồn cầu)",
        "PIC": "Nguyễn Văn A",
        "Ngày bắt đầu": d(-30),
        "Deadline": d(-26),
        "Trạng thái": "Done",
        "Tiến độ (%)": 100,
        "Link Sheet Chi tiết": "→ Sheet 'GD1_Concept'",
    },
    {
        "Mã Task": "GD1-02",
        "Giai đoạn": "GĐ1: Lên Concept & Chuẩn bị",
        "Phòng ban": "Content",
        "Hạng mục công việc": "Lên kịch bản outline (video review, seeding post)",
        "PIC": "Trần Thị B",
        "Ngày bắt đầu": d(-28),
        "Deadline": d(-23),
        "Trạng thái": "Doing",
        "Tiến độ (%)": 60,
        "Link Sheet Chi tiết": "→ Sheet 'GD1_Concept'",
    },
    {
        "Mã Task": "GD1-03",
        "Giai đoạn": "GĐ1: Lên Concept & Chuẩn bị",
        "Phòng ban": "KOC/PR",
        "Hạng mục công việc": "List danh sách KOC mục tiêu & gửi brief",
        "PIC": "Lê Văn C",
        "Ngày bắt đầu": d(-27),
        "Deadline": d(-21),
        "Trạng thái": "Doing",
        "Tiến độ (%)": 40,
        "Link Sheet Chi tiết": "→ Sheet 'GD1_Concept'",
    },

    # ── Giai đoạn 2: Sản xuất & Setup (D-20 → D-7) ──
    {
        "Mã Task": "GD2-01",
        "Giai đoạn": "GĐ2: Sản xuất & Setup",
        "Phòng ban": "Design",
        "Hạng mục công việc": "Chụp ảnh nền trắng / concept cho 2 dòng sản phẩm",
        "PIC": "Phạm Thị D",
        "Ngày bắt đầu": d(-20),
        "Deadline": d(-15),
        "Trạng thái": "To-do",
        "Tiến độ (%)": 0,
        "Link Sheet Chi tiết": "→ Sheet 'GD2_SanXuat'",
    },
    {
        "Mã Task": "GD2-02",
        "Giai đoạn": "GĐ2: Sản xuất & Setup",
        "Phòng ban": "Design",
        "Hạng mục công việc": "Thiết kế Banner sàn (Shopee, Lazada, TikTok Shop)",
        "PIC": "Phạm Thị D",
        "Ngày bắt đầu": d(-16),
        "Deadline": d(-11),
        "Trạng thái": "To-do",
        "Tiến độ (%)": 0,
        "Link Sheet Chi tiết": "→ Sheet 'GD2_SanXuat'",
    },
    {
        "Mã Task": "GD2-03",
        "Giai đoạn": "GĐ2: Sản xuất & Setup",
        "Phòng ban": "Content",
        "Hạng mục công việc": "Viết listing chuẩn SEO (tiêu đề, mô tả, bullet points)",
        "PIC": "Trần Thị B",
        "Ngày bắt đầu": d(-15),
        "Deadline": d(-10),
        "Trạng thái": "To-do",
        "Tiến độ (%)": 0,
        "Link Sheet Chi tiết": "→ Sheet 'GD2_SanXuat'",
    },
    {
        "Mã Task": "GD2-04",
        "Giai đoạn": "GĐ2: Sản xuất & Setup",
        "Phòng ban": "E-commerce",
        "Hạng mục công việc": "Đăng sản phẩm lên sàn (Shopee, Lazada, TikTok Shop)",
        "PIC": "Hoàng Văn E",
        "Ngày bắt đầu": d(-10),
        "Deadline": d(-7),
        "Trạng thái": "To-do",
        "Tiến độ (%)": 0,
        "Link Sheet Chi tiết": "→ Sheet 'GD2_SanXuat'",
    },

    # ── Giai đoạn 3: Teasing & Affiliate (D-7 → D-1) ──
    {
        "Mã Task": "GD3-01",
        "Giai đoạn": "GĐ3: Teasing & Affiliate",
        "Phòng ban": "Booking",
        "Hạng mục công việc": "Gửi PR Box cho KOC (unboxing & review)",
        "PIC": "Lê Văn C",
        "Ngày bắt đầu": d(-7),
        "Deadline": d(-5),
        "Trạng thái": "To-do",
        "Tiến độ (%)": 0,
        "Link Sheet Chi tiết": "→ Sheet 'GD3_Teasing'",
    },
    {
        "Mã Task": "GD3-02",
        "Giai đoạn": "GĐ3: Teasing & Affiliate",
        "Phòng ban": "Creative",
        "Hạng mục công việc": "Lên video teaser trên TikTok (countdown, sneak peek)",
        "PIC": "Trần Thị B",
        "Ngày bắt đầu": d(-6),
        "Deadline": d(-2),
        "Trạng thái": "To-do",
        "Tiến độ (%)": 0,
        "Link Sheet Chi tiết": "→ Sheet 'GD3_Teasing'",
    },
    {
        "Mã Task": "GD3-03",
        "Giai đoạn": "GĐ3: Teasing & Affiliate",
        "Phòng ban": "Vận hành",
        "Hạng mục công việc": "Thiết lập Voucher / Flash Sale trên các sàn",
        "PIC": "Hoàng Văn E",
        "Ngày bắt đầu": d(-5),
        "Deadline": d(-2),
        "Trạng thái": "To-do",
        "Tiến độ (%)": 0,
        "Link Sheet Chi tiết": "→ Sheet 'GD3_Teasing'",
    },
    {
        "Mã Task": "GD3-04",
        "Giai đoạn": "GĐ3: Teasing & Affiliate",
        "Phòng ban": "CSKH",
        "Hạng mục công việc": "Chuẩn bị FAQ & kịch bản trả lời cho CSKH",
        "PIC": "Võ Thị F",
        "Ngày bắt đầu": d(-4),
        "Deadline": d(-1),
        "Trạng thái": "To-do",
        "Tiến độ (%)": 0,
        "Link Sheet Chi tiết": "→ Sheet 'GD3_Teasing'",
    },

    # ── Giai đoạn 4: D-Day Mở bán bùng nổ (D-Day) ──
    {
        "Mã Task": "GD4-01",
        "Giai đoạn": "GĐ4: D-Day Mở bán bùng nổ",
        "Phòng ban": "Creative",
        "Hạng mục công việc": "Đồng loạt push video lên TikTok, Reels, YouTube Shorts",
        "PIC": "Trần Thị B",
        "Ngày bắt đầu": d(0),
        "Deadline": d(0),
        "Trạng thái": "To-do",
        "Tiến độ (%)": 0,
        "Link Sheet Chi tiết": "→ Sheet 'GD4_DDay'",
    },
    {
        "Mã Task": "GD4-02",
        "Giai đoạn": "GĐ4: D-Day Mở bán bùng nổ",
        "Phòng ban": "Booking",
        "Hạng mục công việc": "Push KOC lên bài kèm link Affiliate đồng loạt",
        "PIC": "Lê Văn C",
        "Ngày bắt đầu": d(0),
        "Deadline": d(0),
        "Trạng thái": "To-do",
        "Tiến độ (%)": 0,
        "Link Sheet Chi tiết": "→ Sheet 'GD4_DDay'",
    },
    {
        "Mã Task": "GD4-03",
        "Giai đoạn": "GĐ4: D-Day Mở bán bùng nổ",
        "Phòng ban": "Vận hành",
        "Hạng mục công việc": "Bật TikTok Ads / Shopee Ads chuyển đổi",
        "PIC": "Nguyễn Văn A",
        "Ngày bắt đầu": d(0),
        "Deadline": d(0),
        "Trạng thái": "To-do",
        "Tiến độ (%)": 0,
        "Link Sheet Chi tiết": "→ Sheet 'GD4_DDay'",
    },
    {
        "Mã Task": "GD4-04",
        "Giai đoạn": "GĐ4: D-Day Mở bán bùng nổ",
        "Phòng ban": "CSKH",
        "Hạng mục công việc": "Trực chat và chốt đơn liên tục (Shopee, Lazada, TikTok)",
        "PIC": "Võ Thị F",
        "Ngày bắt đầu": d(0),
        "Deadline": d(0),
        "Trạng thái": "To-do",
        "Tiến độ (%)": 0,
        "Link Sheet Chi tiết": "→ Sheet 'GD4_DDay'",
    },

    # ── Giai đoạn 5: Tối ưu giữa chiến dịch (D+1 → D+3) ──
    {
        "Mã Task": "GD5-01",
        "Giai đoạn": "GĐ5: Tối ưu giữa chiến dịch",
        "Phòng ban": "Vận hành",
        "Hạng mục công việc": "Phân tích số liệu ROAS để điều chỉnh ngân sách Ads",
        "PIC": "Nguyễn Văn A",
        "Ngày bắt đầu": d(1),
        "Deadline": d(2),
        "Trạng thái": "To-do",
        "Tiến độ (%)": 0,
        "Link Sheet Chi tiết": "→ Sheet 'GD5_ToiUu'",
    },
    {
        "Mã Task": "GD5-02",
        "Giai đoạn": "GĐ5: Tối ưu giữa chiến dịch",
        "Phòng ban": "CSKH",
        "Hạng mục công việc": "Xin đánh giá 5 sao từ khách mua sớm (follow-up)",
        "PIC": "Võ Thị F",
        "Ngày bắt đầu": d(1),
        "Deadline": d(3),
        "Trạng thái": "To-do",
        "Tiến độ (%)": 0,
        "Link Sheet Chi tiết": "→ Sheet 'GD5_ToiUu'",
    },
    {
        "Mã Task": "GD5-03",
        "Giai đoạn": "GĐ5: Tối ưu giữa chiến dịch",
        "Phòng ban": "Content",
        "Hạng mục công việc": "Đẩy thêm bài seeding trên group Facebook",
        "PIC": "Trần Thị B",
        "Ngày bắt đầu": d(1),
        "Deadline": d(3),
        "Trạng thái": "To-do",
        "Tiến độ (%)": 0,
        "Link Sheet Chi tiết": "→ Sheet 'GD5_ToiUu'",
    },

    # ── Giai đoạn 6: Đánh giá & Duy trì (D+4 trở đi) ──
    {
        "Mã Task": "GD6-01",
        "Giai đoạn": "GĐ6: Đánh giá & Duy trì",
        "Phòng ban": "Booking",
        "Hạng mục công việc": "Báo cáo đối soát doanh thu KOC (theo từng KOC)",
        "PIC": "Lê Văn C",
        "Ngày bắt đầu": d(4),
        "Deadline": d(7),
        "Trạng thái": "To-do",
        "Tiến độ (%)": 0,
        "Link Sheet Chi tiết": "→ Sheet 'GD6_DanhGia'",
    },
    {
        "Mã Task": "GD6-02",
        "Giai đoạn": "GĐ6: Đánh giá & Duy trì",
        "Phòng ban": "Vận hành",
        "Hạng mục công việc": "Tính lương thưởng KPI 3P (Performance, Productivity, Profit)",
        "PIC": "Hoàng Văn E",
        "Ngày bắt đầu": d(4),
        "Deadline": d(7),
        "Trạng thái": "To-do",
        "Tiến độ (%)": 0,
        "Link Sheet Chi tiết": "→ Sheet 'GD6_DanhGia'",
    },
    {
        "Mã Task": "GD6-03",
        "Giai đoạn": "GĐ6: Đánh giá & Duy trì",
        "Phòng ban": "Marketing",
        "Hạng mục công việc": "Đề xuất chiến lược duy trì bán hàng dài hạn",
        "PIC": "Nguyễn Văn A",
        "Ngày bắt đầu": d(5),
        "Deadline": d(10),
        "Trạng thái": "To-do",
        "Tiến độ (%)": 0,
        "Link Sheet Chi tiết": "→ Sheet 'GD6_DanhGia'",
    },
]

# ============================================================
# 1b. Định nghĩa Team & KPI cho sheet phòng ban
# ============================================================

CREATIVE_DEPTS = {"Creative", "Content", "Design"}
BOOKING_DEPTS = {"Booking", "KOC/PR"}
VANHANH_DEPTS = {"Vận hành", "E-commerce"}
CSKH_DEPTS = {"CSKH"}

# KPI gợi ý theo từng task
KPI_MAP = {
    # Team Creative
    "GD1-02": "Hoàn thành 3 kịch bản outline được duyệt",
    "GD2-01": "Bộ 20 ảnh nền trắng + 10 ảnh concept / dòng SP",
    "GD2-02": "6 banner (2 sàn x 3 size) đúng spec, duyệt lần 1",
    "GD2-03": "Listing đạt điểm SEO ≥ 80 trên tool check",
    "GD3-02": "3 video teaser đạt ≥ 50K views trong 48h",
    "GD4-01": "Tổng views D-Day ≥ 200K, engagement rate ≥ 5%",
    "GD5-03": "Đăng ≥ 10 bài seeding, reach ≥ 30K",
    # Team Booking
    "GD1-03": "List ≥ 15 KOC phù hợp, gửi brief trong 3 ngày",
    "GD3-01": "100% PR Box gửi đúng hạn, có tracking",
    "GD4-02": "≥ 10 KOC đăng bài đúng D-Day, kèm link Affiliate",
    "GD6-01": "Báo cáo đối soát hoàn thành trong 3 ngày, sai lệch < 2%",
    # Team Vận Hành
    "GD2-04": "100% sản phẩm lên sàn đúng hạn, listing đầy đủ hình + mô tả",
    "GD3-03": "Voucher/Flash Sale active trước D-Day 48h, test thành công",
    "GD4-03": "ROAS Ads D-Day ≥ 3.0, budget burn rate đúng kế hoạch",
    "GD5-01": "Báo cáo ROAS trong 24h, đề xuất điều chỉnh ngân sách cụ thể",
    "GD6-02": "Bảng tính KPI 3P hoàn thành, sai số < 1%",
    # Team CSKH
    "GD3-04": "FAQ ≥ 30 câu, kịch bản chat đạt duyệt trước D-2",
    "GD4-04": "Tỷ lệ phản hồi < 5 phút, tỷ lệ chốt đơn ≥ 40%",
    "GD5-02": "≥ 50 đánh giá 5 sao trong 3 ngày đầu",
}

# ============================================================
# 2. Tạo DataFrame và ghi ra Excel
# ============================================================

df = pd.DataFrame(tasks)

# Tạo DataFrame cho Team Creative & Team Booking
df_creative = df[df["Phòng ban"].isin(CREATIVE_DEPTS)].copy()
df_creative["KPI / Yêu cầu đầu ra"] = df_creative["Mã Task"].map(KPI_MAP).fillna("")
df_creative["Ghi chú nội bộ"] = ""

df_booking = df[df["Phòng ban"].isin(BOOKING_DEPTS)].copy()
df_booking["KPI / Yêu cầu đầu ra"] = df_booking["Mã Task"].map(KPI_MAP).fillna("")
df_booking["Ghi chú nội bộ"] = ""

df_vanhanh = df[df["Phòng ban"].isin(VANHANH_DEPTS)].copy()
df_vanhanh["KPI / Yêu cầu đầu ra"] = df_vanhanh["Mã Task"].map(KPI_MAP).fillna("")
df_vanhanh["Ghi chú nội bộ"] = ""

df_cskh = df[df["Phòng ban"].isin(CSKH_DEPTS)].copy()
df_cskh["KPI / Yêu cầu đầu ra"] = df_cskh["Mã Task"].map(KPI_MAP).fillna("")
df_cskh["Ghi chú nội bộ"] = ""

# ============================================================
# 1c. Dữ liệu KPI Tracker (Lương & Thưởng)
# ============================================================

kpi_tracker_data = [
    {
        "Vị trí": "Vận hành sàn (E-commerce)",
        "Lương cứng": 10_000_000,
        "KPI Trách nhiệm tối đa": 5_000_000,
        "Cơ chế Thưởng Doanh số": "Từ mốc 1 tỷ: 0.5% DS; 2 tỷ: 0.8% DS; 3 tỷ+: 1% DS",
        "Kết quả đánh giá %": 85,
        "Thưởng dự kiến": 4_250_000,
    },
    {
        "Vị trí": "Creative (Content + Design)",
        "Lương cứng": 9_000_000,
        "KPI Trách nhiệm tối đa": 4_000_000,
        "Cơ chế Thưởng Doanh số": "Thưởng cố định theo campaign: 3tr/chiến dịch đạt KPI",
        "Kết quả đánh giá %": 90,
        "Thưởng dự kiến": 3_600_000,
    },
    {
        "Vị trí": "Booking (KOC/PR)",
        "Lương cứng": 8_000_000,
        "KPI Trách nhiệm tối đa": 4_000_000,
        "Cơ chế Thưởng Doanh số": "2% doanh thu từ link Affiliate KOC quản lý",
        "Kết quả đánh giá %": 75,
        "Thưởng dự kiến": 3_000_000,
    },
    {
        "Vị trí": "CSKH (Chăm sóc khách hàng)",
        "Lương cứng": 7_500_000,
        "KPI Trách nhiệm tối đa": 3_000_000,
        "Cơ chế Thưởng Doanh số": "Thưởng chốt đơn: 5K/đơn chat; 10K/đơn livestream",
        "Kết quả đánh giá %": 80,
        "Thưởng dự kiến": 2_400_000,
    },
    {
        "Vị trí": "Marketing (Chiến lược)",
        "Lương cứng": 12_000_000,
        "KPI Trách nhiệm tối đa": 6_000_000,
        "Cơ chế Thưởng Doanh số": "Thưởng theo ROAS: đạt ≥ 3.0 được 5tr; ≥ 5.0 được 8tr",
        "Kết quả đánh giá %": 92,
        "Thưởng dự kiến": 5_520_000,
    },
]

df_kpi = pd.DataFrame(kpi_tracker_data)

OUTPUT_FILE = "TIAN_Launch_MasterPlan_6Phases.xlsx"

with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
    df.to_excel(writer, sheet_name="Master Dashboard", index=False)
    df_creative.to_excel(writer, sheet_name="Team_Creative", index=False)
    df_booking.to_excel(writer, sheet_name="Team_Booking", index=False)
    df_vanhanh.to_excel(writer, sheet_name="Team_VanHanh", index=False)
    df_cskh.to_excel(writer, sheet_name="Team_CSKH", index=False)
    df_kpi.to_excel(writer, sheet_name="KPI_Tracker", index=False)

# ============================================================
# 3. Định dạng (format) bằng openpyxl
# ============================================================

wb = load_workbook(OUTPUT_FILE)
ws = wb["Master Dashboard"]

# --- 3a. Style cho header ---
header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

for col_idx in range(1, ws.max_column + 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# --- 3b. Style cho data rows ---
data_font = Font(name="Arial", size=10)
data_alignment = Alignment(vertical="center", wrap_text=True)

# Màu nền cho trạng thái
status_colors = {
    "To-do": PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),   # Xanh nhạt
    "Doing": PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid"),    # Vàng nhạt
    "Done": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),     # Xanh lá nhạt
    "Overdue": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),  # Đỏ nhạt
}

STATUS_COL = 8  # Cột "Trạng thái" (H)
PROGRESS_COL = 9  # Cột "Tiến độ (%)" (I)

for row_idx in range(2, ws.max_row + 1):
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = thin_border

    # Tô màu cột Trạng thái
    status_cell = ws.cell(row=row_idx, column=STATUS_COL)
    status_val = status_cell.value
    if status_val in status_colors:
        status_cell.fill = status_colors[status_val]

    # Format cột Tiến độ (%)
    progress_cell = ws.cell(row=row_idx, column=PROGRESS_COL)
    progress_cell.number_format = '0"%"'
    progress_cell.alignment = Alignment(horizontal="center", vertical="center")

# --- 3c. Data Validation cho cột Trạng thái ---
dv = DataValidation(
    type="list",
    formula1='"To-do,Doing,Done,Overdue"',
    allow_blank=True,
    showDropDown=False,
)
dv.prompt = "Chọn trạng thái"
dv.promptTitle = "Trạng thái"
ws.add_data_validation(dv)
dv.add(f"H2:H1000")

# --- 3d. Điều chỉnh độ rộng cột ---
column_widths = {
    "A": 10,   # Mã Task
    "B": 32,   # Giai đoạn
    "C": 14,   # Phòng ban
    "D": 50,   # Hạng mục công việc
    "E": 16,   # PIC
    "F": 14,   # Ngày bắt đầu
    "G": 14,   # Deadline
    "H": 12,   # Trạng thái
    "I": 12,   # Tiến độ (%)
    "J": 22,   # Link Sheet Chi tiết
}

for col_letter, width in column_widths.items():
    ws.column_dimensions[col_letter].width = width

# --- 3e. Freeze header row ---
ws.freeze_panes = "A2"

# --- 3f. Tô màu xen kẽ nhóm Giai đoạn ---
odd_phase_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")   # Xám nhạt (GĐ1, GĐ3)
even_phase_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")   # Trắng   (GĐ2, GĐ4)

for row_idx in range(2, ws.max_row + 1):
    giai_doan = ws.cell(row=row_idx, column=2).value or ""
    is_odd_phase = any(tag in giai_doan for tag in ["GĐ1", "GĐ3", "GĐ5"])
    fill = odd_phase_fill if is_odd_phase else even_phase_fill
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        # Không ghi đè màu trạng thái
        if col_idx != STATUS_COL:
            cell.fill = fill

# ============================================================
# 4. Tạo Hyperlink tại cột "Link Sheet Chi tiết" trên Master Dashboard
# ============================================================

LINK_COL = 10  # Cột J
hyperlink_font = Font(name="Arial", size=10, color="0563C1", underline="single")

# Xây dựng map: Mã Task → (sheet_name, row_in_sheet)
task_to_sheet_row = {}
for sheet_name, dept_set in [("Team_Creative", CREATIVE_DEPTS), ("Team_Booking", BOOKING_DEPTS),
                              ("Team_VanHanh", VANHANH_DEPTS), ("Team_CSKH", CSKH_DEPTS)]:
    team_ws = wb[sheet_name]
    for row_idx in range(2, team_ws.max_row + 1):
        ma_task = team_ws.cell(row=row_idx, column=1).value
        task_to_sheet_row[ma_task] = (sheet_name, row_idx)

for row_idx in range(2, ws.max_row + 1):
    ma_task = ws.cell(row=row_idx, column=1).value
    cell = ws.cell(row=row_idx, column=LINK_COL)

    if ma_task in task_to_sheet_row:
        sheet_name, target_row = task_to_sheet_row[ma_task]
        cell.value = f"→ {sheet_name}"
        cell.hyperlink = Hyperlink(ref=cell.coordinate, location=f"{sheet_name}!A{target_row}", display=cell.value)
        cell.font = hyperlink_font
    # Các task không thuộc Creative/Booking giữ nguyên text gốc

# ============================================================
# 5. Định dạng các sheet Team
# ============================================================

TEAM_HEADER_FILLS = {
    "Team_Creative": PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid"),   # Xanh dương
    "Team_Booking": PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid"),    # Cam
    "Team_VanHanh": PatternFill(start_color="548235", end_color="548235", fill_type="solid"),    # Xanh lá đậm
    "Team_CSKH": PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid"),       # Tím
}

# Độ rộng cột cho team sheets (10 cột gốc + 2 cột mới)
team_col_widths = {
    "A": 10, "B": 32, "C": 14, "D": 50, "E": 16,
    "F": 14, "G": 14, "H": 12, "I": 12, "J": 22,
    "K": 42, "L": 28,
}

for sheet_name in ["Team_Creative", "Team_Booking", "Team_VanHanh", "Team_CSKH"]:
    team_ws = wb[sheet_name]
    team_header_fill = TEAM_HEADER_FILLS[sheet_name]

    # Header
    for col_idx in range(1, team_ws.max_column + 1):
        cell = team_ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = team_header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Data rows
    for row_idx in range(2, team_ws.max_row + 1):
        for col_idx in range(1, team_ws.max_column + 1):
            cell = team_ws.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border

        # Tô màu trạng thái
        status_cell = team_ws.cell(row=row_idx, column=STATUS_COL)
        if status_cell.value in status_colors:
            status_cell.fill = status_colors[status_cell.value]

        # Format tiến độ
        progress_cell = team_ws.cell(row=row_idx, column=PROGRESS_COL)
        progress_cell.number_format = '0"%"'
        progress_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data Validation trạng thái
    team_dv = DataValidation(
        type="list",
        formula1='"To-do,Doing,Done,Overdue"',
        allow_blank=True,
        showDropDown=False,
    )
    team_ws.add_data_validation(team_dv)
    team_dv.add("H2:H1000")

    # Độ rộng cột
    for col_letter, width in team_col_widths.items():
        team_ws.column_dimensions[col_letter].width = width

    # Freeze header
    team_ws.freeze_panes = "A2"

    # Link quay về Master Dashboard tại cột J
    back_font = Font(name="Arial", size=10, color="0563C1", underline="single")
    for row_idx in range(2, team_ws.max_row + 1):
        cell = team_ws.cell(row=row_idx, column=LINK_COL)
        cell.value = "← Master Dashboard"
        cell.hyperlink = Hyperlink(ref=cell.coordinate, location="'Master Dashboard'!A1", display=cell.value)
        cell.font = back_font

# ============================================================
# 6. Định dạng sheet KPI_Tracker
# ============================================================

kpi_ws = wb["KPI_Tracker"]
kpi_header_fill = PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid")  # Vàng đậm

# Header
for col_idx in range(1, kpi_ws.max_column + 1):
    cell = kpi_ws.cell(row=1, column=col_idx)
    cell.font = header_font
    cell.fill = kpi_header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# Data rows
for row_idx in range(2, kpi_ws.max_row + 1):
    for col_idx in range(1, kpi_ws.max_column + 1):
        cell = kpi_ws.cell(row=row_idx, column=col_idx)
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = thin_border

    # Format tiền tệ: Lương cứng (B), KPI max (C), Thưởng dự kiến (F)
    for money_col in [2, 3, 6]:
        cell = kpi_ws.cell(row=row_idx, column=money_col)
        cell.number_format = '#,##0'
        cell.alignment = Alignment(horizontal="right", vertical="center")

    # Format % cho Kết quả đánh giá (E)
    eval_cell = kpi_ws.cell(row=row_idx, column=5)
    eval_cell.number_format = '0"%"'
    eval_cell.alignment = Alignment(horizontal="center", vertical="center")

# Độ rộng cột KPI_Tracker
kpi_col_widths = {
    "A": 30,   # Vị trí
    "B": 16,   # Lương cứng
    "C": 22,   # KPI Trách nhiệm tối đa
    "D": 55,   # Cơ chế Thưởng Doanh số
    "E": 18,   # Kết quả đánh giá %
    "F": 18,   # Thưởng dự kiến
}
for col_letter, width in kpi_col_widths.items():
    kpi_ws.column_dimensions[col_letter].width = width

kpi_ws.freeze_panes = "A2"

# Công thức tính Thưởng dự kiến = KPI max × Kết quả %
for row_idx in range(2, kpi_ws.max_row + 1):
    cell = kpi_ws.cell(row=row_idx, column=6)
    cell.value = f"=C{row_idx}*E{row_idx}/100"
    cell.number_format = '#,##0'

# Link quay về Master Dashboard
for row_idx in range(2, kpi_ws.max_row + 1):
    # Thêm cột G làm link
    cell = kpi_ws.cell(row=row_idx, column=7)
    cell.value = "← Master Dashboard"
    cell.hyperlink = Hyperlink(ref=cell.coordinate, location="'Master Dashboard'!A1", display=cell.value)
    cell.font = Font(name="Arial", size=10, color="0563C1", underline="single")
# Header cho cột G
g_header = kpi_ws.cell(row=1, column=7)
g_header.value = "Điều hướng"
g_header.font = header_font
g_header.fill = kpi_header_fill
g_header.alignment = header_alignment
g_header.border = thin_border
kpi_ws.column_dimensions["G"].width = 22

# ============================================================
# 7. Conditional Formatting cho TẤT CẢ sheets có cột Trạng thái
# ============================================================

# Các sheet có cột Trạng thái (H) và Tiến độ (I)
task_sheets = ["Master Dashboard", "Team_Creative", "Team_Booking", "Team_VanHanh", "Team_CSKH"]

# Màu cho Conditional Formatting (cập nhật theo yêu cầu mới)
cf_status_styles = {
    "To-do":   (PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),   # Xám
                Font(name="Arial", size=10, color="595959")),
    "Doing":   (PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid"),    # Vàng
                Font(name="Arial", size=10, color="7F6003")),
    "Done":    (PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),    # Xanh lá
                Font(name="Arial", size=10, color="006100")),
    "Overdue": (PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),    # Đỏ
                Font(name="Arial", size=10, bold=True, color="9C0006")),
}

for sheet_name in task_sheets:
    target_ws = wb[sheet_name]
    max_row = target_ws.max_row
    status_range = f"H2:H{max(max_row, 100)}"
    progress_range = f"I2:I{max(max_row, 100)}"

    # Conditional Formatting cho cột Trạng thái
    for status_val, (cf_fill, cf_font) in cf_status_styles.items():
        target_ws.conditional_formatting.add(
            status_range,
            CellIsRule(
                operator="equal",
                formula=[f'"{status_val}"'],
                fill=cf_fill,
                font=cf_font,
            ),
        )

    # Data Bars (thanh dữ liệu xanh) cho cột Tiến độ (%)
    target_ws.conditional_formatting.add(
        progress_range,
        DataBarRule(
            start_type="num",
            start_value=0,
            end_type="num",
            end_value=100,
            color="4472C4",   # Xanh dương
        ),
    )

# Cũng cập nhật lại màu trạng thái tĩnh trên Master Dashboard theo màu mới
updated_status_colors = {
    "To-do":   PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
    "Doing":   PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid"),
    "Done":    PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "Overdue": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
}
for sheet_name in task_sheets:
    target_ws = wb[sheet_name]
    for row_idx in range(2, target_ws.max_row + 1):
        status_cell = target_ws.cell(row=row_idx, column=STATUS_COL)
        if status_cell.value in updated_status_colors:
            status_cell.fill = updated_status_colors[status_cell.value]

# Data Bar cho cột Kết quả đánh giá trên KPI_Tracker (cột E)
kpi_ws.conditional_formatting.add(
    f"E2:E{kpi_ws.max_row}",
    DataBarRule(
        start_type="num",
        start_value=0,
        end_type="num",
        end_value=100,
        color="70AD47",   # Xanh lá
    ),
)

# ============================================================
# 8. Lưu file hoàn chỉnh
# ============================================================

wb.save(OUTPUT_FILE)
print(f"✅ File '{OUTPUT_FILE}' đã được tạo thành công!")
print(f"   - Sheets: {wb.sheetnames}")
print(f"   - Tổng số task: {len(tasks)}")
print(f"   - Giai đoạn 1: 3 tasks (D-30 → D-21) — Lên Concept & Chuẩn bị")
print(f"   - Giai đoạn 2: 4 tasks (D-20 → D-7)  — Sản xuất & Setup")
print(f"   - Giai đoạn 3: 4 tasks (D-7 → D-1)   — Teasing & Affiliate")
print(f"   - Giai đoạn 4: 4 tasks (D-Day)        — Mở bán bùng nổ")
print(f"   - Giai đoạn 5: 3 tasks (D+1 → D+3)   — Tối ưu giữa chiến dịch")
print(f"   - Giai đoạn 6: 3 tasks (D+4 → D+10)  — Đánh giá & Duy trì")
print(f"   - Team_Creative: {len(df_creative)} tasks")
print(f"   - Team_Booking:  {len(df_booking)} tasks")
print(f"   - Team_VanHanh:  {len(df_vanhanh)} tasks")
print(f"   - Team_CSKH:     {len(df_cskh)} tasks")
print(f"   - KPI_Tracker:   {len(df_kpi)} vị trí")
print(f"   - Conditional Formatting: Trạng thái (4 màu) + Data Bars (Tiến độ)")
print(f"   - D-Day giả định: {D_DAY.strftime('%Y-%m-%d')}")
